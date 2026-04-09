from __future__ import annotations

import json
import os
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2 import service_account
from openpyxl import load_workbook


GOOGLE_DRIVE_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
GOOGLE_SHEETS_SCOPE = "https://www.googleapis.com/auth/spreadsheets.readonly"
GOOGLE_CALENDAR_SCOPE = "https://www.googleapis.com/auth/calendar.readonly"
GOOGLE_DRIVE_API_BASE = "https://www.googleapis.com/drive/v3/files"
GOOGLE_SHEETS_API_BASE = "https://sheets.googleapis.com/v4/spreadsheets"
GOOGLE_CALENDAR_API_BASE = "https://www.googleapis.com/calendar/v3/calendars"
GOOGLE_DOC_MIME = "application/vnd.google-apps.document"


@dataclass
class DataSourceConfig:
    source_type: str
    workspace_dir: Path
    google_service_account_json: str
    drive_readme_file_id: str
    drive_line_log_file_id: str
    drive_application_tracker_file_id: str
    sheets_revenue_spreadsheet_id: str
    sheets_revenue_range: str
    google_calendar_id: str
    google_calendar_timezone: str


def load_data_source_config() -> DataSourceConfig:
    load_dotenv()
    return DataSourceConfig(
        source_type=os.getenv("PROJECT_DATA_SOURCE", "local"),
        workspace_dir=Path(
            os.getenv(
                "WORKSPACE_DIR",
                str(Path(__file__).resolve().parent),
            )
        ),
        google_service_account_json=os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", ""),
        drive_readme_file_id=os.getenv("GOOGLE_DRIVE_README_FILE_ID", ""),
        drive_line_log_file_id=os.getenv("GOOGLE_DRIVE_LINE_LOG_FILE_ID", ""),
        drive_application_tracker_file_id=os.getenv(
            "GOOGLE_DRIVE_APPLICATION_TRACKER_FILE_ID", ""
        ),
        sheets_revenue_spreadsheet_id=os.getenv(
            "GOOGLE_SHEETS_REVENUE_SPREADSHEET_ID", ""
        ),
        sheets_revenue_range=os.getenv(
            "GOOGLE_SHEETS_REVENUE_RANGE", "Summary!A1:Z50"
        ),
        google_calendar_id=os.getenv("GOOGLE_CALENDAR_ID", ""),
        google_calendar_timezone=os.getenv(
            "GOOGLE_CALENDAR_TIMEZONE", "Asia/Tokyo"
        ),
    )


def build_project_snapshot(max_line_log_lines: int = 220) -> dict[str, Any]:
    config = load_data_source_config()
    if config.source_type == "google_workspace":
        return build_google_workspace_snapshot(config, max_line_log_lines)
    return build_local_snapshot(config.workspace_dir, max_line_log_lines)


def build_local_snapshot(workspace_dir: Path, max_line_log_lines: int = 220) -> dict[str, Any]:
    return {
        "data_source": "local",
        "workspace_dir": str(workspace_dir),
        "project_readme": read_local_readme(workspace_dir),
        "recent_line_log": read_local_line_log(workspace_dir, max_line_log_lines),
        "application_tracker": read_local_application_tracker(workspace_dir),
        "revenue_sheet_url": read_local_revenue_sheet_link(workspace_dir),
        "revenue_sheet_preview": None,
        "calendar_today": [],
    }


def build_google_workspace_snapshot(
    config: DataSourceConfig,
    max_line_log_lines: int = 220,
) -> dict[str, Any]:
    client = GoogleWorkspaceClient.from_service_account_json(
        config.google_service_account_json
    )
    warnings: list[str] = []

    readme_text = safe_read(
        warnings,
        "README",
        lambda: client.read_drive_text(config.drive_readme_file_id)
        if config.drive_readme_file_id
        else "",
    )
    line_log_text = safe_read(
        warnings,
        "LINE_LOG",
        lambda: client.read_drive_text(config.drive_line_log_file_id)
        if config.drive_line_log_file_id
        else "",
    )
    if line_log_text:
        trimmed = [line for line in line_log_text.splitlines() if line.strip()]
        line_log_text = "\n".join(trimmed[-max_line_log_lines:])

    application_tracker = safe_read(
        warnings,
        "APPLICATION_TRACKER",
        lambda: client.read_application_tracker_xlsx(
            config.drive_application_tracker_file_id
        )
        if config.drive_application_tracker_file_id
        else {"rows": [], "status_summary": {}},
    )

    revenue_sheet_preview = safe_read(
        warnings,
        "REVENUE_SHEET",
        lambda: client.read_sheet_values(
            config.sheets_revenue_spreadsheet_id,
            config.sheets_revenue_range,
        )
        if config.sheets_revenue_spreadsheet_id
        else None,
    )

    revenue_sheet_url = (
        f"https://docs.google.com/spreadsheets/d/{config.sheets_revenue_spreadsheet_id}/edit"
        if config.sheets_revenue_spreadsheet_id
        else ""
    )
    calendar_today = safe_read(
        warnings,
        "GOOGLE_CALENDAR",
        lambda: client.read_calendar_events_today(
            config.google_calendar_id,
            config.google_calendar_timezone,
        )
        if config.google_calendar_id
        else [],
    )

    return {
        "data_source": "google_workspace",
        "workspace_dir": None,
        "project_readme": readme_text,
        "recent_line_log": line_log_text,
        "application_tracker": application_tracker,
        "revenue_sheet_url": revenue_sheet_url,
        "revenue_sheet_preview": revenue_sheet_preview,
        "calendar_today": calendar_today,
        "data_warnings": warnings,
    }


def safe_read(warnings: list[str], label: str, fn) -> Any:
    try:
        return fn()
    except Exception as exc:
        warnings.append(f"{label}: {exc.__class__.__name__}: {exc}")
        return default_value_for(label)


def default_value_for(label: str) -> Any:
    if label == "APPLICATION_TRACKER":
        return {"rows": [], "status_summary": {}}
    if label == "GOOGLE_CALENDAR":
        return []
    return "" if label in {"README", "LINE_LOG"} else None


def read_local_readme(workspace_dir: Path) -> str:
    path = workspace_dir / "README.md"
    return path.read_text(encoding="utf-8") if path.exists() else ""


def read_local_line_log(workspace_dir: Path, max_lines: int = 220) -> str:
    candidates = sorted(workspace_dir.glob("[[]LINE[]]*.txt"))
    if not candidates:
        return ""

    path = candidates[0]
    lines = path.read_text(encoding="utf-8").splitlines()
    trimmed = [line for line in lines if line.strip()][-max_lines:]
    return "\n".join(trimmed)


def read_local_application_tracker(workspace_dir: Path) -> dict[str, Any]:
    tracker_path = workspace_dir / "申請管理.xlsx"
    if not tracker_path.exists():
        return {"rows": [], "status_summary": {}}
    with tracker_path.open("rb") as f:
        return parse_application_tracker_workbook(f.read())


def read_local_revenue_sheet_link(workspace_dir: Path) -> str:
    gsheet_path = workspace_dir / "キッチンカー収支.gsheet"
    if not gsheet_path.exists():
        return ""

    try:
        data = json.loads(gsheet_path.read_text(encoding="utf-8"))
        doc_id = data.get("doc_id", "")
    except json.JSONDecodeError:
        return ""

    if not doc_id:
        return ""
    return f"https://docs.google.com/spreadsheets/d/{doc_id}/edit"


def parse_application_tracker_workbook(raw_bytes: bytes) -> dict[str, Any]:
    workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True)
    sheet = workbook["申請一覧"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]

    items: list[dict[str, Any]] = []
    status_summary: dict[str, int] = {}

    for row in rows[1:]:
        if not any(row):
            continue
        item = {
            headers[idx]: row[idx]
            for idx in range(min(len(headers), len(row)))
            if headers[idx]
        }
        items.append(item)
        status = str(item.get("ステータス", "")).strip() or "未設定"
        status_summary[status] = status_summary.get(status, 0) + 1

    return {"rows": items, "status_summary": status_summary}


class GoogleWorkspaceClient:
    def __init__(self, credentials: service_account.Credentials) -> None:
        self.credentials = credentials

    @classmethod
    def from_service_account_json(cls, raw_json: str) -> "GoogleWorkspaceClient":
        if not raw_json:
            raise RuntimeError("Missing GOOGLE_SERVICE_ACCOUNT_JSON")
        info = json.loads(raw_json)
        credentials = service_account.Credentials.from_service_account_info(
            info,
            scopes=[GOOGLE_DRIVE_SCOPE, GOOGLE_SHEETS_SCOPE, GOOGLE_CALENDAR_SCOPE],
        )
        return cls(credentials)

    def _authorized_headers(self) -> dict[str, str]:
        if not self.credentials.valid:
            self.credentials.refresh(GoogleAuthRequest())
        return {"Authorization": f"Bearer {self.credentials.token}"}

    def _get_drive_metadata(self, file_id: str) -> dict[str, Any]:
        response = requests.get(
            f"{GOOGLE_DRIVE_API_BASE}/{file_id}",
            headers=self._authorized_headers(),
            params={"fields": "id,name,mimeType"},
            timeout=30,
        )
        response.raise_for_status()
        return response.json()

    def read_drive_text(self, file_id: str) -> str:
        metadata = self._get_drive_metadata(file_id)
        mime_type = metadata.get("mimeType", "")

        if mime_type == GOOGLE_DOC_MIME:
            response = requests.get(
                f"{GOOGLE_DRIVE_API_BASE}/{file_id}/export",
                headers=self._authorized_headers(),
                params={"mimeType": "text/plain"},
                timeout=30,
            )
        else:
            response = requests.get(
                f"{GOOGLE_DRIVE_API_BASE}/{file_id}",
                headers=self._authorized_headers(),
                params={"alt": "media"},
                timeout=30,
            )
        response.raise_for_status()
        return response.text

    def read_application_tracker_xlsx(self, file_id: str) -> dict[str, Any]:
        response = requests.get(
            f"{GOOGLE_DRIVE_API_BASE}/{file_id}",
            headers=self._authorized_headers(),
            params={"alt": "media"},
            timeout=30,
        )
        response.raise_for_status()
        return parse_application_tracker_workbook(response.content)

    def read_sheet_values(self, spreadsheet_id: str, cell_range: str) -> dict[str, Any]:
        encoded_range = quote(cell_range, safe="")
        response = requests.get(
            f"{GOOGLE_SHEETS_API_BASE}/{spreadsheet_id}/values/{encoded_range}",
            headers=self._authorized_headers(),
            timeout=30,
        )
        response.raise_for_status()
        return response.json()

    def read_calendar_events_today(
        self,
        calendar_id: str,
        timezone_name: str,
    ) -> list[dict[str, Any]]:
        tz = ZoneInfo(timezone_name)
        now = datetime.now(tz)
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day.replace(hour=23, minute=59, second=59)
        encoded_calendar_id = quote(calendar_id, safe="")
        response = requests.get(
            f"{GOOGLE_CALENDAR_API_BASE}/{encoded_calendar_id}/events",
            headers=self._authorized_headers(),
            params={
                "singleEvents": "true",
                "orderBy": "startTime",
                "timeMin": start_of_day.isoformat(),
                "timeMax": end_of_day.isoformat(),
                "timeZone": timezone_name,
                "fields": (
                    "items(id,summary,description,location,"
                    "start(dateTime,date),end(dateTime,date),status,htmlLink)"
                ),
            },
            timeout=30,
        )
        response.raise_for_status()
        items = response.json().get("items", [])
        return [
            {
                "id": item.get("id"),
                "summary": item.get("summary") or "(No title)",
                "description": item.get("description"),
                "location": item.get("location"),
                "status": item.get("status"),
                "start": item.get("start", {}),
                "end": item.get("end", {}),
                "htmlLink": item.get("htmlLink"),
            }
            for item in items
            if item.get("status") != "cancelled"
        ]
