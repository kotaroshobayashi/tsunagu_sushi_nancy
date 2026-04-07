#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


JST = ZoneInfo("Asia/Tokyo")
LINE_PUSH_URL = "https://api.line.me/v2/bot/message/push"
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
MAX_LINE_TEXT_LENGTH = 5000


@dataclass
class Settings:
    workspace_dir: Path
    gemini_api_key: str
    gemini_model: str
    line_channel_access_token: str
    line_target_id: str
    report_day: int
    report_hour: int
    report_minute: int
    timezone: ZoneInfo


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a weekly project report from local files and push it to LINE."
    )
    parser.add_argument(
        "--run-once",
        action="store_true",
        help="Generate and send one report immediately.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Generate the report and print it without sending to LINE.",
    )
    return parser.parse_args()


def load_settings() -> Settings:
    load_dotenv()

    workspace_dir = Path(
        os.getenv(
            "WORKSPACE_DIR",
            "/Users/kotaroshobayashi/Library/CloudStorage/GoogleDrive-shobayashi.kotaro@gmail.com/My Drive/TSUNAGU/SushiBiz/Nancy",
        )
    )
    gemini_api_key = require_env("GEMINI_API_KEY")
    line_channel_access_token = require_env("LINE_CHANNEL_ACCESS_TOKEN")
    line_target_id = require_env("LINE_TARGET_ID")

    return Settings(
        workspace_dir=workspace_dir,
        gemini_api_key=gemini_api_key,
        gemini_model=os.getenv("GEMINI_MODEL", "gemini-2.0-flash"),
        line_channel_access_token=line_channel_access_token,
        line_target_id=line_target_id,
        report_day=int(os.getenv("REPORT_DAY", "0")),  # Monday
        report_hour=int(os.getenv("REPORT_HOUR", "22")),
        report_minute=int(os.getenv("REPORT_MINUTE", "0")),
        timezone=ZoneInfo(os.getenv("REPORT_TIMEZONE", "Asia/Tokyo")),
    )


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def read_readme(workspace_dir: Path) -> str:
    path = workspace_dir / "README.md"
    return path.read_text(encoding="utf-8") if path.exists() else ""


def read_line_log(workspace_dir: Path, max_lines: int = 220) -> str:
    candidates = sorted(workspace_dir.glob("[[]LINE[]]*.txt"))
    if not candidates:
        return ""

    path = candidates[0]
    lines = path.read_text(encoding="utf-8").splitlines()
    trimmed = [line for line in lines if line.strip()][-max_lines:]
    return "\n".join(trimmed)


def read_application_tracker(workspace_dir: Path) -> dict[str, Any]:
    tracker_path = workspace_dir / "申請管理.xlsx"
    if not tracker_path.exists():
        return {"rows": [], "status_summary": {}}

    workbook = load_workbook(tracker_path, data_only=True)
    sheet = workbook["申請一覧"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]

    items: list[dict[str, Any]] = []
    status_summary: dict[str, int] = {}

    for row in rows[1:]:
        if not any(row):
            continue
        item = {
            headers[idx]: row[idx]
            for idx in range(min(len(headers), len(row)))
            if headers[idx]
        }
        items.append(item)
        status = str(item.get("ステータス", "")).strip() or "未設定"
        status_summary[status] = status_summary.get(status, 0) + 1

    return {"rows": items, "status_summary": status_summary}


def read_revenue_sheet_link(workspace_dir: Path) -> str:
    gsheet_path = workspace_dir / "キッチンカー収支.gsheet"
    if not gsheet_path.exists():
        return ""

    try:
        data = json.loads(gsheet_path.read_text(encoding="utf-8"))
        doc_id = data.get("doc_id", "")
    except json.JSONDecodeError:
        return ""

    if not doc_id:
        return ""
    return f"https://docs.google.com/spreadsheets/d/{doc_id}/edit"


def build_report_input(settings: Settings) -> dict[str, Any]:
    tracker = read_application_tracker(settings.workspace_dir)
    now = datetime.now(settings.timezone)

    return {
        "generated_at": now.isoformat(),
        "workspace_dir": str(settings.workspace_dir),
        "project_readme": read_readme(settings.workspace_dir),
        "recent_line_log": read_line_log(settings.workspace_dir),
        "application_tracker": tracker,
        "revenue_sheet_url": read_revenue_sheet_link(settings.workspace_dir),
    }


def generate_weekly_report(settings: Settings, source_data: dict[str, Any]) -> str:
    instructions = (
        "You are an operations assistant for a Japanese/French food event project. "
        "Create a concise weekly report in Japanese for a LINE group. "
        "Use exact dates where possible. "
        "Focus on current status, application progress, blockers, and next actions. "
        "Do not invent facts. If something is unclear, say that it is not yet confirmed. "
        "Keep the report readable in LINE and under 3500 Japanese characters. "
        "Use this structure exactly:\n"
        "【今週の要点】\n"
        "【申請・書類の進捗】\n"
        "【未解決事項】\n"
        "【来週のアクション】\n"
        "【担当別メモ】"
    )

    prompt = (
        "以下のローカルプロジェクト情報を元に、週次レポートを作ってください。\n\n"
        f"{json.dumps(source_data, ensure_ascii=False, indent=2)}"
    )
    return generate_with_gemini(
        api_key=settings.gemini_api_key,
        model=settings.gemini_model,
        system_instruction=instructions,
        user_prompt=prompt,
    )


def send_line_message(settings: Settings, text: str) -> None:
    headers = {
        "Authorization": f"Bearer {settings.line_channel_access_token}",
        "Content-Type": "application/json",
    }

    chunks = split_for_line(text)
    for chunk in chunks:
        payload = {
            "to": settings.line_target_id,
            "messages": [{"type": "text", "text": chunk}],
        }
        retry_key = str(uuid.uuid4())
        response = requests.post(
            LINE_PUSH_URL,
            headers={**headers, "X-Line-Retry-Key": retry_key},
            json=payload,
            timeout=30,
        )
        response.raise_for_status()


def generate_with_gemini(
    *,
    api_key: str,
    model: str,
    system_instruction: str,
    user_prompt: str,
) -> str:
    url = GEMINI_API_URL.format(model=model)
    payload = {
        "systemInstruction": {
            "parts": [{"text": system_instruction}],
        },
        "contents": [
            {
                "role": "user",
                "parts": [{"text": user_prompt}],
            }
        ],
        "generationConfig": {
            "temperature": 0.4,
            "topP": 0.9,
            "maxOutputTokens": 2200,
        },
    }
    response = requests.post(
        url,
        params={"key": api_key},
        headers={"Content-Type": "application/json"},
        json=payload,
        timeout=60,
    )
    response.raise_for_status()
    data = response.json()

    candidates = data.get("candidates") or []
    if not candidates:
        raise RuntimeError(f"Gemini returned no candidates: {data}")

    parts = candidates[0].get("content", {}).get("parts", [])
    text = "".join(part.get("text", "") for part in parts).strip()
    if not text:
        raise RuntimeError(f"Gemini returned empty text: {data}")
    return text


def split_for_line(text: str) -> list[str]:
    if len(text) <= MAX_LINE_TEXT_LENGTH:
        return [text]

    chunks: list[str] = []
    current = ""
    for line in text.splitlines(keepends=True):
        if len(current) + len(line) > MAX_LINE_TEXT_LENGTH:
            if current:
                chunks.append(current.rstrip())
            current = line
        else:
            current += line

    if current.strip():
        chunks.append(current.rstrip())
    return chunks


def next_run_at(settings: Settings, now: datetime | None = None) -> datetime:
    now = now or datetime.now(settings.timezone)
    candidate = now.replace(
        hour=settings.report_hour,
        minute=settings.report_minute,
        second=0,
        microsecond=0,
    )

    days_ahead = (settings.report_day - candidate.weekday()) % 7
    candidate = candidate + timedelta(days=days_ahead)
    if candidate <= now:
        candidate = candidate + timedelta(days=7)
    return candidate


def sleep_until(target: datetime) -> None:
    while True:
        now = datetime.now(target.tzinfo)
        remaining = (target - now).total_seconds()
        if remaining <= 0:
            return
        time.sleep(min(remaining, 60))


def run_once(settings: Settings, dry_run: bool) -> None:
    source_data = build_report_input(settings)
    report = generate_weekly_report(settings, source_data)

    if dry_run:
        print(report)
        return

    send_line_message(settings, report)
    logging.info("Weekly report sent to LINE.")


def run_scheduler(settings: Settings, dry_run: bool) -> None:
    while True:
        target = next_run_at(settings)
        logging.info("Next weekly report scheduled for %s", target.isoformat())
        sleep_until(target)
        try:
            run_once(settings, dry_run=dry_run)
        except Exception as exc:
            logging.exception("Failed to send weekly report: %s", exc)


def main() -> int:
    args = parse_args()
    configure_logging()

    try:
        settings = load_settings()
        if args.run_once:
            run_once(settings, dry_run=args.dry_run)
        else:
            run_scheduler(settings, dry_run=args.dry_run)
    except KeyboardInterrupt:
        logging.info("Stopped by user.")
        return 130
    except Exception as exc:
        logging.exception("Bot stopped with error: %s", exc)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
